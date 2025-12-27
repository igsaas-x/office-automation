import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ...domain.entities.check_in import CheckIn
from ...domain.entities.employee import Employee
from ...domain.entities.group import Group


class ExcelExportService:
    """Service for exporting check-in reports to Excel format"""

    def __init__(self):
        self.exports_dir = "exports/checkins"
        self._ensure_export_directory()

    def _ensure_export_directory(self):
        """Create export directory if it doesn't exist"""
        os.makedirs(self.exports_dir, exist_ok=True)

    def generate_checkin_report(
        self,
        check_ins: List[CheckIn],
        employees: Dict[int, Employee],
        group: Group,
        month: int,
        year: int
    ) -> str:
        """
        Generate Excel report for check-ins

        Args:
            check_ins: List of check-in records
            employees: Dictionary mapping employee_id to Employee
            group: Group entity
            month: Month number (1-12)
            year: Year (e.g., 2024)

        Returns:
            Absolute file path to generated Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Check-In Report - {month:02d}/{year}"

        # Add summary section
        self._add_summary_section(ws, group, month, year, check_ins, employees)

        # Add table headers
        self._add_table_headers(ws)

        # Populate data rows
        self._populate_data_rows(ws, check_ins, employees)

        # Apply formatting
        self._apply_formatting(ws)

        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"checkin_report_{group.id}_{timestamp}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)

        return os.path.abspath(filepath)

    def _add_summary_section(
        self,
        ws,
        group: Group,
        month: int,
        year: int,
        check_ins: List[CheckIn],
        employees: Dict[int, Employee]
    ):
        """Add summary section at the top of the report"""
        month_names = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]

        # Business name
        business_name = group.business_name or group.name
        ws['A1'] = f"Business Name: {business_name}"
        ws['A1'].font = Font(bold=True, size=14)

        # Report period
        ws['A2'] = f"Report Period: {month_names[month-1]} {year}"
        ws['A2'].font = Font(bold=True, size=12)

        # Total employees
        unique_employees = set(ci.employee_id for ci in check_ins)
        ws['A3'] = f"Total Employees: {len(unique_employees)}"
        ws['A3'].font = Font(size=11)

        # Total check-ins
        ws['A4'] = f"Total Check-ins: {len(check_ins)}"
        ws['A4'].font = Font(size=11)

        # Blank row
        # Row 5 is blank

    def _add_table_headers(self, ws):
        """Add table headers at row 6"""
        headers = ['Employee', 'Date', 'Time', 'Location']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Freeze header row
        ws.freeze_panes = 'A7'

    def _populate_data_rows(
        self,
        ws,
        check_ins: List[CheckIn],
        employees: Dict[int, Employee]
    ):
        """Populate data rows sorted by date and employee name"""
        # Sort check-ins by timestamp (date), then by employee name
        sorted_check_ins = sorted(
            check_ins,
            key=lambda ci: (
                ci.timestamp.date(),
                employees.get(ci.employee_id, Employee(
                    id=None,
                    telegram_id="",
                    name="Unknown",
                    phone=None,
                    role=None,
                    date_start_work=None,
                    probation_months=None,
                    base_salary=None,
                    bonus=None,
                    created_at=datetime.now()
                )).name
            )
        )

        # Populate rows starting from row 7
        row_num = 7
        for check_in in sorted_check_ins:
            employee = employees.get(check_in.employee_id)
            employee_name = employee.name if employee else "Unknown"

            # Employee name
            ws.cell(row=row_num, column=1, value=employee_name)

            # Date (DD/MM/YYYY)
            date_str = check_in.timestamp.strftime("%d/%m/%Y")
            ws.cell(row=row_num, column=2, value=date_str)

            # Time (HH:MM)
            time_str = check_in.timestamp.strftime("%H:%M")
            ws.cell(row=row_num, column=3, value=time_str)

            # Location (Google Maps hyperlink or N/A)
            if check_in.location and check_in.location.latitude and check_in.location.longitude:
                maps_url = f"https://www.google.com/maps?q={check_in.location.latitude},{check_in.location.longitude}"
                cell = ws.cell(row=row_num, column=4, value="View Map")
                cell.hyperlink = maps_url
                cell.font = Font(color="0563C1", underline="single")
            else:
                ws.cell(row=row_num, column=4, value="N/A")

            row_num += 1

    def _apply_formatting(self, ws):
        """Apply formatting to the worksheet"""
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Employee
        ws.column_dimensions['B'].width = 15  # Date
        ws.column_dimensions['C'].width = 10  # Time
        ws.column_dimensions['D'].width = 15  # Location

        # Apply alternating row colors (starting from row 7)
        max_row = ws.max_row
        for row_num in range(7, max_row + 1):
            if (row_num - 7) % 2 == 1:  # Alternate rows
                for col_num in range(1, 5):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Center align date and time columns
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')